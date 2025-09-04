#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
АСИНХРОННЫЙ скрипт для выявления товаров без изображений в Compo PIM API
"""

import requests
import asyncio
import aiohttp
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

class AsyncCompoImageChecker:
    def __init__(self):
        self.token = None
        self.base_url = "https://pim.uroven.pro/api/v1"
        self.headers = {"Content-Type": "application/json"}
        
    def authenticate(self):
        """Получение токена авторизации"""
        response = requests.post(
            f"{self.base_url}/sign-in/",
            json={
                "login": "s.andrey",
                "password": "KZh-4g2-YFx-Jgm",
                "remember": True
            },
            headers=self.headers
        )
        response.raise_for_status()
        self.token = response.json()["data"]["access"]["token"]
        self.headers["Authorization"] = f"Bearer {self.token}"
    
    async def get_products_without_images(self):
        """Асинхронное получение всех товаров без изображений"""
        scroll_id = None
        products_without_images = []
        batch_num = 0
        total_processed = 0
        
        # Создаем сессию для параллельных запросов
        connector = aiohttp.TCPConnector(limit=20, limit_per_host=10)
        timeout = aiohttp.ClientTimeout(total=10)
        
        async with aiohttp.ClientSession(connector=connector, timeout=timeout) as session:
            while True:
                batch_num += 1
                url = f"{self.base_url}/product/scroll"
                if scroll_id:
                    url += f"?scrollId={scroll_id}"
                
                # API запрос делаем синхронно (авторизация)
                response = requests.get(url, headers=self.headers)
                response.raise_for_status()
                response_data = response.json()
                
                if not response_data.get("success", False):
                    print(f"❌ API вернул ошибку: {response_data.get('message', 'Неизвестная ошибка')}")
                    break
                
                data = response_data.get("data", {})
                current_batch = data.get("productElasticDtos", [])
                total_products = data.get("total", "неизвестно")
                
                if not current_batch:
                    print("⛔ Нет товаров в пакете, завершаем...")
                    break
                
                # Параллельно проверяем все товары в пакете
                tasks = []
                for product in current_batch:
                    tasks.append(self.check_product_async(product))
                
                results = await asyncio.gather(*tasks)
                
                batch_without_images = 0
                for product, has_images in zip(current_batch, results):
                    total_processed += 1
                    if not has_images:
                        batch_without_images += 1
                        products_without_images.append({
                            "id": product.get("id"),
                            "code_1c": product.get("articul", ""),
                            "header": product.get("header", "")
                        })
                
                if batch_num % 10 == 0 or batch_without_images > 0:  # Показываем прогресс
                    print(f"🚀 Пакет {batch_num}: {total_processed}/{total_products} товаров | Без изображений: {len(products_without_images)}")
                
                scroll_id = data.get("scrollId")
                if not scroll_id:
                    print("⛔ Нет scrollId, завершаем...")
                    break
                    
        return products_without_images
    
    async def check_product_async(self, product):
        """Асинхронная проверка наличия изображений у товара"""
        # Простая проверка полей - не требует сетевых запросов
        await asyncio.sleep(0)  # Небольшая задержка для асинхронности
        
        picture = product.get("picture")
        pictures = product.get("pictures", [])
        
        has_main_picture = bool(picture)
        has_additional_pictures = bool(pictures and len(pictures) > 0)
        
        return has_main_picture or has_additional_pictures
    
    def save_to_excel(self, products):
        """Сохранение результатов в Excel файл"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Товары без изображений"
        
        headers = ["ID товара", "Код 1С", "Название товара", "Статус"]
        
        # Записываем заголовки
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Записываем данные
        for row, product in enumerate(products, 2):
            ws.cell(row=row, column=1, value=product["id"])
            ws.cell(row=row, column=2, value=product["code_1c"])
            ws.cell(row=row, column=3, value=product["header"])
            ws.cell(row=row, column=4, value="Нет изображений")
        
        # Автоподбор ширины колонок
        for col in range(1, len(headers) + 1):
            max_length = 0
            column = get_column_letter(col)
            
            for row in range(1, len(products) + 2):
                try:
                    if len(str(ws[f'{column}{row}'].value)) > max_length:
                        max_length = len(str(ws[f'{column}{row}'].value))
                except:
                    pass
            
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
        
        # Добавляем информацию о дате формирования отчета
        ws.cell(row=len(products) + 3, column=1, value=f"Отчет сформирован: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        ws.cell(row=len(products) + 4, column=1, value=f"Всего товаров без изображений: {len(products)}")
        
        # Сохраняем файл
        filename = f"products_without_images_ASYNC_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        
        print(f"\n✅ Результат сохранен в Excel файл: {filename}")
        print(f"📊 Файл содержит {len(products)} товаров без изображений")
        
        return len(products)

async def main_async():
    checker = AsyncCompoImageChecker()
    
    try:
        print("🔐 Авторизация...")
        checker.authenticate()
        print("✅ Авторизация успешна")
        
        print("🚀 Асинхронное получение товаров...")
        products = await checker.get_products_without_images()
        
        print(f"\n📊 Итого найдено: {len(products)} товаров без изображений")
        
        checker.save_to_excel(products)
        print(f"✅ Успешно: найдено {len(products)} товаров без изображений")
        return True
        
    except Exception as e:
        print(f"❌ Ошибка: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def main():
    return asyncio.run(main_async())

if __name__ == "__main__":
    main()
