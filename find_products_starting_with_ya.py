#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Асинхронный скрипт для поиска товаров с проблемными названиями (начинаются с "я" или "Я", но не "Ящик")
и сохранения их в Excel для контентщиков
"""

import os
import asyncio
from datetime import datetime
from supabase import create_client
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

load_dotenv()

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")


async def find_problem_products():
    """Поиск товаров с проблемными названиями (начинаются с "я"/"Я", но не "Ящик")"""
    try:
        client = create_client(SUPABASE_URL, SUPABASE_KEY)
        print("✅ Подключение к Supabase установлено")
        
        loop = asyncio.get_event_loop()
        
        # Получаем товары, начинающиеся с "Я" (большая буква)
        response1 = await loop.run_in_executor(
            None,
            lambda: client.table("onec_catalog")
                .select("product_name, code_1c, article")
                .ilike("product_name", "Я%")
                .execute()
        )
        
        # Получаем товары, начинающиеся с "я" (маленькая буква)
        response2 = await loop.run_in_executor(
            None,
            lambda: client.table("onec_catalog")
                .select("product_name, code_1c, article")
                .ilike("product_name", "я%")
                .execute()
        )
        
        # Объединяем результаты
        all_products = (response1.data or []) + (response2.data or [])
        
        # Убираем дубликаты по code_1c
        seen_codes = set()
        unique_products = []
        for product in all_products:
            code = product.get("code_1c")
            if code and code not in seen_codes:
                seen_codes.add(code)
                unique_products.append(product)
        
        # Фильтруем: исключаем товары с "Ящик" в названии
        problem_products = []
        for product in unique_products:
            name = product.get("product_name", "")
            if name and "Ящик" not in name:
                problem_products.append({
                    "product_name": name,
                    "code_1c": product.get("code_1c"),
                    "article": product.get("article")
                })
        
        print(f"\n📊 Найдено проблемных товаров: {len(problem_products)}")
        
        return problem_products
        
    except Exception as e:
        print(f"❌ Ошибка: {e}")
        import traceback
        traceback.print_exc()
        return []


def save_to_excel(products):
    """Сохранение результатов в Excel файл"""
    if not products:
        print("❌ Нет товаров для сохранения")
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Проблемные товары"
    
    headers = ["Название", "Код 1С", "Артикул"]
    
    # Записываем заголовки
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="4F81BD", end_color="4F81BD", fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center")
    
    # Записываем данные
    for row, product in enumerate(products, 2):
        ws.cell(row=row, column=1, value=product["product_name"])
        ws.cell(row=row, column=2, value=product["code_1c"])
        ws.cell(row=row, column=3, value=product["article"])
    
    # Автоподбор ширины колонок
    for col in range(1, len(headers) + 1):
        max_length = 0
        column = get_column_letter(col)
        
        for row in range(1, len(products) + 2):
            try:
                cell_value = str(ws[f"{column}{row}"].value or "")
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except:
                pass
        
        adjusted_width = min((max_length + 2) * 1.2, 80)
        ws.column_dimensions[column].width = adjusted_width
    
    # Добавляем информацию о дате
    ws.cell(
        row=len(products) + 3,
        column=1,
        value=f"Отчет сформирован: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    )
    ws.cell(
        row=len(products) + 4,
        column=1,
        value=f"Всего проблемных товаров: {len(products)}"
    )
    
    # Сохраняем файл
    filename = f"problem_products_ya_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    
    print(f"\n✅ Результат сохранен в Excel файл: {filename}")
    print(f"📊 Файл содержит {len(products)} проблемных товаров")


async def main():
    products = await find_problem_products()
    if products:
        save_to_excel(products)
    else:
        print("✅ Проблемных товаров не найдено")


if __name__ == "__main__":
    asyncio.run(main())

