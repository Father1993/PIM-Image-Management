#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Скрипт для выявления товаров с изображениями шириной менее 500px в Compo PIM API
"""

import requests
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from PIL import Image
from io import BytesIO
import logging

class SmallImageChecker:
    def __init__(self):
        self.token = None
        self.base_url = "https://pim.uroven.pro/api/v1"
        self.headers = {"Content-Type": "application/json"}
        self.image_cache = {}  # Кэш для размеров изображений
        logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    def authenticate(self):
        """Получение токена авторизации"""
        response = requests.post(
            f"{self.base_url}/sign-in/",
            json={
                "login": "s.andrey",
                "password": "KZh-4g2-YFx-Jgm",
                "remember": True
            },
            headers=self.headers
        )
        response.raise_for_status()
        self.token = response.json()["data"]["access"]["token"]
        self.headers["Authorization"] = f"Bearer {self.token}"
    
    def get_products_with_small_images(self):
        """Получение товаров с изображениями шириной менее 500px"""
        scroll_id = None
        products_with_small_images = []
        batch_num = 0
        total_processed = 0
        
        while True:
            batch_num += 1
            url = f"{self.base_url}/product/scroll"
            if scroll_id:
                url += f"?scrollId={scroll_id}"
            
            response = requests.get(url, headers=self.headers)
            response.raise_for_status()
            response_data = response.json()
            
            if not response_data.get("success", False):
                print(f"❌ API вернул ошибку: {response_data.get('message', 'Неизвестная ошибка')}")
                break
            
            data = response_data.get("data", {})
            current_batch = data.get("productElasticDtos", [])
            total_products = data.get("total", "неизвестно")
            
            if not current_batch:
                print("⛔ Нет товаров в пакете, завершаем...")
                break
            
            batch_small_images = 0
            
            for product in current_batch:
                total_processed += 1
                product_id = product.get("id")
                
                # Проверяем размеры изображений
                small_images = self.check_image_sizes(product)
                
                if small_images:
                    batch_small_images += 1
                    products_with_small_images.append({
                        "id": product_id,
                        "code_1c": product.get("articul", ""),
                        "header": product.get("header", ""),
                        "small_images": small_images
                    })
            
            if batch_num % 10 == 0 or batch_small_images > 0:  # Выводим каждый 10-й пакет или при находках
                print(f"📊 Пакет {batch_num}: {total_processed}/{total_products} товаров | Найдено: {len(products_with_small_images)}")
            
            scroll_id = data.get("scrollId")
            if not scroll_id:
                print("⛔ Нет scrollId, завершаем...")
                break
                
        return products_with_small_images
    
    def check_image_sizes(self, product):
        """Проверка размеров изображений товара"""
        small_images = []
        
        # Проверяем основное изображение
        picture = product.get("picture")
        if picture:
            width, height = self.get_image_info(picture)
            if width and width < 500:
                small_images.append(f"Основное: {picture} ({width}px)")
        
        # Проверяем дополнительные изображения
        pictures = product.get("pictures", [])
        for pic in pictures:
            width, height = self.get_image_info(pic)
            if width and width < 500:
                small_images.append(f"Доп.: {pic} ({width}px)")
        
        return small_images
    
    def get_image_info(self, image_name):
        """Получение размера изображения (оптимизированно с кэшем)"""
        if image_name in self.image_cache:
            return self.image_cache[image_name]
        
        try:
            image_url = f"https://pim.uroven.pro/pictures/originals/{image_name}"
            
            # Загружаем только первые 2KB для получения размеров
            headers = {"Range": "bytes=0-2047"}
            response = requests.get(image_url, headers=headers, timeout=3)
            
            if response.status_code in [200, 206]:  # 206 для частичного контента
                image = Image.open(BytesIO(response.content))
                width, height = image.width, image.height
                self.image_cache[image_name] = (width, height)
                return width, height
                        
        except Exception:
            pass
        
        # Кэшируем неудачные попытки как None
        self.image_cache[image_name] = (None, None)
        return None, None
    
    def save_to_excel(self, products):
        """Сохранение результатов в Excel файл"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Товары с маленькими изображениями"
        
        headers = ["ID товара", "Код 1С", "Название товара", "Маленькие изображения"]
        
        # Записываем заголовки
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Записываем данные
        for row, product in enumerate(products, 2):
            ws.cell(row=row, column=1, value=product["id"])
            ws.cell(row=row, column=2, value=product["code_1c"])
            ws.cell(row=row, column=3, value=product["header"])
            ws.cell(row=row, column=4, value="; ".join(product["small_images"]))
        
        # Автоподбор ширины колонок
        for col in range(1, len(headers) + 1):
            max_length = 0
            column = get_column_letter(col)
            
            for row in range(1, len(products) + 2):
                try:
                    cell_value = str(ws[f'{column}{row}'].value)
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except:
                    pass
            
            adjusted_width = min((max_length + 2) * 1.2, 50)  # Ограничиваем максимальную ширину
            ws.column_dimensions[column].width = adjusted_width
        
        # Добавляем информацию о дате формирования отчета
        ws.cell(row=len(products) + 3, column=1, value=f"Отчет сформирован: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        ws.cell(row=len(products) + 4, column=1, value=f"Всего товаров с маленькими изображениями: {len(products)}")
        
        # Сохраняем файл
        filename = f"products_small_images_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        
        print(f"\n✅ Результат сохранен в Excel файл: {filename}")
        print(f"📊 Файл содержит {len(products)} товаров с маленькими изображениями")
        
        return len(products)

def main():
    checker = SmallImageChecker()
    
    try:
        print("🔐 Авторизация...")
        checker.authenticate()
        print("✅ Авторизация успешна")
        
        print("🔍 Проверка размеров изображений...")
        products = checker.get_products_with_small_images()
        
        print(f"\n📊 Итого найдено: {len(products)} товаров с изображениями < 500px")
        print(f"💾 Кэш изображений: {len(checker.image_cache)} уникальных файлов")
        
        if products:
            checker.save_to_excel(products)
        else:
            print("Товары с маленькими изображениями не найдены")
        
        return True
        
    except Exception as e:
        print(f"❌ Ошибка: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    main()
