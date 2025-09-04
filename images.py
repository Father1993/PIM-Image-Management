#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Оптимизированный скрипт для выявления товаров без изображений в Compo PIM API
"""

import requests
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

class CompoImageChecker:
    def __init__(self):
        self.token = None
        # Используем единый URL для API
        self.base_url = "https://pim.uroven.pro/api/v1"
        self.headers = {"Content-Type": "application/json"}
        
    def authenticate(self):
        """Получение токена авторизации"""
        response = requests.post(
            f"{self.base_url}/sign-in/",
            json={
                "login": "s.andrey",
                "password": "KZh-4g2-YFx-Jgm",
                "remember": True
            },
            headers=self.headers
        )
        response.raise_for_status()
        self.token = response.json()["data"]["access"]["token"]
        self.headers["Authorization"] = f"Bearer {self.token}"
    
    def get_products_without_images(self):
        """Получение всех товаров без изображений через scroll API"""
        scroll_id = None
        products_without_images = []
        batch_num = 0
        total_processed = 0
        
        while True:
            batch_num += 1
            # Формируем URL с параметрами
            url = f"{self.base_url}/product/scroll"
            if scroll_id:
                url += f"?scrollId={scroll_id}"
            
            # Получаем пакет товаров
            response = requests.get(url, headers=self.headers)
            response.raise_for_status()
            response_data = response.json()
            
            # Проверяем успешность запроса
            if not response_data.get("success", False):
                print(f"❌ API вернул ошибку: {response_data.get('message', 'Неизвестная ошибка')}")
                break
            
            data = response_data.get("data", {})
            current_batch = data.get("productElasticDtos", [])  # Исправлено название поля!
            total_products = data.get("total", "неизвестно")
            
            if not current_batch:
                print("⛔ Нет товаров в пакете, завершаем...")
                break
            
            batch_without_images = 0
            
            # Проверяем товары из текущего пакета
            for product in current_batch:
                total_processed += 1
                product_id = product.get("id")
                has_images = self.has_images(product)
                
                if not has_images:
                    batch_without_images += 1
                    products_without_images.append({
                        "id": product_id,
                        "code_1c": product.get("syncUid", ""),
                        "header": product.get("header", "")
                    })
            
            print(f"🔍 Пакет {batch_num}: обработано {len(current_batch)} товаров, найдено {batch_without_images} без изображений")
            print(f"📊 Прогресс: {total_processed}/{total_products} товаров | Без изображений: {len(products_without_images)}")
            
            # Проверяем, нужно ли продолжать
            scroll_id = data.get("scrollId")
            
            if not scroll_id:
                print("⛔ Нет scrollId, завершаем...")
                break
                
        return products_without_images
    
    def has_images(self, product):
        """Проверка наличия изображений у товара"""
        picture = product.get("picture")
        pictures = product.get("pictures", [])
        
        # Проверяем основное изображение (строка с именем файла)
        has_main_picture = bool(picture)
        
        # Проверяем дополнительные изображения (массив)
        has_additional_pictures = bool(pictures and len(pictures) > 0)
        
        return has_main_picture or has_additional_pictures
    
    def save_to_excel(self, products):
        """Сохранение результатов в Excel файл"""
        # Создаем новую рабочую книгу и выбираем активный лист
        wb = Workbook()
        ws = wb.active
        ws.title = "Товары без изображений"
        
        # Заголовки колонок
        headers = ["ID товара", "Код 1С", "Название товара", "Статус"]
        
        # Записываем заголовки
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            # Стилизуем заголовки
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Записываем данные
        for row, product in enumerate(products, 2):
            ws.cell(row=row, column=1, value=product["id"])
            ws.cell(row=row, column=2, value=product["code_1c"])
            ws.cell(row=row, column=3, value=product["header"])
            ws.cell(row=row, column=4, value="Нет изображений")
        
        # Автоподбор ширины колонок
        for col in range(1, len(headers) + 1):
            max_length = 0
            column = get_column_letter(col)
            
            for row in range(1, len(products) + 2):
                try:
                    if len(str(ws[f'{column}{row}'].value)) > max_length:
                        max_length = len(str(ws[f'{column}{row}'].value))
                except:
                    pass
            
            # Добавляем немного пространства для красоты
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
        
        # Добавляем информацию о дате формирования отчета
        ws.cell(row=len(products) + 3, column=1, value=f"Отчет сформирован: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        ws.cell(row=len(products) + 4, column=1, value=f"Всего товаров без изображений: {len(products)}")
        
        # Сохраняем файл
        filename = f"products_without_images_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        
        print(f"\n✅ Результат сохранен в Excel файл: {filename}")
        print(f"📊 Файл содержит {len(products)} товаров без изображений")
        
        return len(products)

def main():
    checker = CompoImageChecker()
    
    try:
        print("🔐 Авторизация...")
        checker.authenticate()
        print("✅ Авторизация успешна")
        
        print("🔍 Получение товаров...")
        products = checker.get_products_without_images()
        
        print(f"\n📊 Итог:")
        print(f"   Товаров без изображений: {len(products)}")
        
        count = checker.save_to_excel(products)
        
        print(f"✅ Успешно: найдено {count} товаров без изображений")
        return True
        
    except Exception as e:
        print(f"❌ Ошибка: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    main()