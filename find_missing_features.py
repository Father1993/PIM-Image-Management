#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""Поиск товаров без характеристик Тип/Вид/Назначение в Compo PIM API."""

import asyncio
import json
import os
import sys
from datetime import datetime

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

PIM_API_URL = os.getenv("PIM_API_URL")
PIM_LOGIN = os.getenv("PIM_LOGIN")
PIM_PASSWORD = os.getenv("PIM_PASSWORD")

AUTH_PAYLOAD = {"login": PIM_LOGIN, "password": PIM_PASSWORD, "remember": True}
REQUIRED_FEATURES = ("Тип", "Вид", "Назначение")


def ensure_config() -> None:
    missing = [
        name
        for name, value in (
            ("PIM_API_URL", PIM_API_URL),
            ("PIM_LOGIN", PIM_LOGIN),
            ("PIM_PASSWORD", PIM_PASSWORD),
        )
        if not value
    ]
    if missing:
        raise SystemExit(
            "Отсутствуют переменные окружения: " + ", ".join(sorted(missing))
        )


def authenticate() -> str:
    response = requests.post(
        f"{PIM_API_URL}/sign-in/", json=AUTH_PAYLOAD, timeout=10
    )
    response.raise_for_status()
    payload = response.json()
    if not payload.get("success"):
        raise RuntimeError(payload.get("message", "Не удалось авторизоваться"))
    try:
        return payload["data"]["access"]["token"]
    except KeyError as exc:
        raise RuntimeError(
            "Не удалось получить токен авторизации из ответа API"
        ) from exc


def fetch_products(token: str):
    headers = {"Authorization": f"Bearer {token}"}
    scroll_id = None
    while True:
        params = {"scrollId": scroll_id} if scroll_id else {}
        response = requests.get(
            f"{PIM_API_URL}/product/scroll/",
            headers=headers,
            params=params,
            timeout=20,
        )
        response.raise_for_status()
        payload = response.json()
        if not payload.get("success"):
            raise RuntimeError(payload.get("message", "Ошибка при получении товаров"))
        data = payload.get("data") or {}
        products = data.get("productElasticDtos") or []
        if not products:
            break
        for product in products:
            yield product
        scroll_id = data.get("scrollId")
        if not scroll_id:
            break


def has_feature_value(product: dict, target_header: str) -> bool:
    for feature in product.get("features") or []:
        if feature.get("header") == target_header:
            values = feature.get("values") or []
            return any((value.get("header") or "").strip() for value in values)
    return False


def analyze_product(product: dict):
    missing = [
        feature_name
        for feature_name in REQUIRED_FEATURES
        if not has_feature_value(product, feature_name)
    ]
    if missing:
        return {
            "id": product.get("id"),
            "header": product.get("header") or "",
            "missing": missing,
        }
    return None


async def collect_missing_products(token: str):
    print("🔄 Загрузка и анализ товаров...")
    missing_products = []
    tasks = []
    processed = 0

    async def flush():
        nonlocal tasks, processed
        if not tasks:
            return
        results = await asyncio.gather(*tasks)
        tasks = []
        processed += len(results)
        for result in results:
            if result:
                missing_products.append(result)
        print(f"   • Обработано {processed} товаров")
    for product in fetch_products(token):
        tasks.append(asyncio.to_thread(analyze_product, product))
        if len(tasks) >= 32:
            await flush()
    await flush()
    return missing_products


def main() -> int:
    ensure_config()
    print("🔄 Авторизация...")
    try:
        token = authenticate()
    except Exception as exc:
        print(f"❌ Авторизация не удалась: {exc}")
        return 1
    print("✅ Авторизация завершена.")

    try:
        missing_products = asyncio.run(collect_missing_products(token))
    except Exception as exc:
        print(f"❌ Ошибка при получении товаров: {exc}")
        return 1

    # Сохранение результата в Excel файл
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"missing_features_{timestamp}.xlsx"
    
    save_to_excel(missing_products, filename)
    
    print(f"✅ Результат сохранен в файл: {filename}")
    print(f"📊 Найдено товаров с неполными характеристиками: {len(missing_products)}")
    
    return 0


def save_to_excel(products, filename):
    """Сохранение результатов в Excel файл"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Товары без характеристик"

    headers = ["ID товара", "Название товара", "Ссылка на PIM", "Отсутствующие характеристики"]

    # Записываем заголовки
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="D32F2F", end_color="D32F2F", fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center")

    # Записываем данные
    for row, product in enumerate(products, 2):
        ws.cell(row=row, column=1, value=product["id"])
        ws.cell(row=row, column=2, value=product["header"])
        ws.cell(row=row, column=3, value=f"https://pim.uroven.pro/cabinet/pim/catalog/21/products/item/edit/{product['id']}")
        ws.cell(row=row, column=4, value=", ".join(product["missing"]))

    # Автоподбор ширины колонок
    for col in range(1, len(headers) + 1):
        max_length = 0
        column = get_column_letter(col)

        for row in range(1, len(products) + 2):
            try:
                cell_value = str(ws[f"{column}{row}"].value)
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except:
                pass

        adjusted_width = min((max_length + 2) * 1.2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Добавляем информацию о дате формирования отчета
    ws.cell(
        row=len(products) + 3,
        column=1,
        value=f"Отчет сформирован: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
    )
    ws.cell(
        row=len(products) + 4,
        column=1,
        value=f"Всего товаров без характеристик: {len(products)}",
    )

    # Сохраняем файл
    wb.save(filename)


if __name__ == "__main__":
    sys.exit(main())