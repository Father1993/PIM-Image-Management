#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
АСИНХРОННЫЙ скрипт для выявления товаров с изображениями шириной более 750px
"""

import requests
import asyncio
import aiohttp
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from PIL import Image
from io import BytesIO
import logging

class AsyncBigImageChecker:
    def __init__(self):
        self.token = None
        self.base_url = "https://pim.uroven.pro/api/v1"
        self.headers = {"Content-Type": "application/json"}
        self.image_cache = {}
        logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    
    def authenticate(self):
        """Получение токена авторизации (синхронно)"""
        response = requests.post(
            f"{self.base_url}/sign-in/",
            json={
                "login": "s.andrey",
                "password": "KZh-4g2-YFx-Jgm",
                "remember": True
            },
            headers=self.headers
        )
        response.raise_for_status()
        self.token = response.json()["data"]["access"]["token"]
        self.headers["Authorization"] = f"Bearer {self.token}"
    
    async def get_products_with_big_images(self):
        """Получение товаров с изображениями шириной более 750px"""
        scroll_id = None
        products_with_big_images = []
        batch_num = 0
        total_processed = 0
        
        # Создаем сессию для изображений
        connector = aiohttp.TCPConnector(limit=50, limit_per_host=20)
        timeout = aiohttp.ClientTimeout(total=5)
        
        async with aiohttp.ClientSession(connector=connector, timeout=timeout) as session:
            while True:
                batch_num += 1
                url = f"{self.base_url}/product/scroll"
                if scroll_id:
                    url += f"?scrollId={scroll_id}"
                
                # API товаров делаем синхронно (авторизация)
                response = requests.get(url, headers=self.headers)
                response.raise_for_status()
                response_data = response.json()
                
                if not response_data.get("success", False):
                    print(f"❌ API вернул ошибку: {response_data.get('message', 'Неизвестная ошибка')}")
                    break
                
                data = response_data.get("data", {})
                current_batch = data.get("productElasticDtos", [])
                total_products = data.get("total", "неизвестно")
                
                if not current_batch:
                    print("⛔ Нет товаров в пакете, завершаем...")
                    break
                
                # Параллельно проверяем все товары в пакете
                tasks = []
                for product in current_batch:
                    tasks.append(self.check_product_images_async(session, product))
                
                results = await asyncio.gather(*tasks)
                
                batch_big_images = 0
                for product, big_images in zip(current_batch, results):
                    total_processed += 1
                    if big_images:
                        batch_big_images += 1
                        products_with_big_images.append({
                            "id": product.get("id"),
                            "code_1c": product.get("articul", ""),
                            "header": product.get("header", ""),
                            "big_images": big_images
                        })
                
                if batch_num % 5 == 0 or batch_big_images > 0:  # Чаще показываем прогресс
                    print(f"🚀 Пакет {batch_num}: {total_processed}/{total_products} товаров | Найдено: {len(products_with_big_images)}")
                
                scroll_id = data.get("scrollId")
                if not scroll_id:
                    print("⛔ Нет scrollId, завершаем...")
                    break
                    
        return products_with_big_images
    
    async def check_product_images_async(self, session, product):
        """Асинхронная проверка размеров изображений товара"""
        tasks = []
        
        # Основное изображение
        picture = product.get("picture")
        if picture:
            tasks.append(self.get_image_info_async(session, picture, "Основное"))
        
        # Дополнительные изображения
        pictures = product.get("pictures", [])
        for pic in pictures:
            tasks.append(self.get_image_info_async(session, pic, "Доп."))
        
        if not tasks:
            return []
        
        # Параллельно проверяем все изображения товара
        results = await asyncio.gather(*tasks)
        return [result for result in results if result]
    
    async def get_image_info_async(self, session, image_name, img_type):
        """Асинхронное получение размера изображения"""
        if image_name in self.image_cache:
            width, height = self.image_cache[image_name]
            if width and width > 750:
                return f"{img_type}: {image_name} ({width}px)"
            return None
        
        try:
            image_url = f"https://pim.uroven.pro/pictures/originals/{image_name}"
            headers = {"Range": "bytes=0-2047"}
            
            async with session.get(image_url, headers=headers) as response:
                if response.status in [200, 206]:
                    content = await response.read()
                    image = Image.open(BytesIO(content))
                    width, height = image.width, image.height
                    self.image_cache[image_name] = (width, height)
                    
                    if width > 750:
                        return f"{img_type}: {image_name} ({width}px)"
                        
        except Exception:
            pass
        
        # Кэшируем неудачные попытки
        self.image_cache[image_name] = (None, None)
        return None
    
    def save_to_excel(self, products):
        """Сохранение результатов в Excel файл"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Товары с большими изображениями"
        
        headers = ["ID товара", "Код 1С", "Название товара", "Большие изображения"]
        
        # Записываем заголовки
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Записываем данные
        for row, product in enumerate(products, 2):
            ws.cell(row=row, column=1, value=product["id"])
            ws.cell(row=row, column=2, value=product["code_1c"])
            ws.cell(row=row, column=3, value=product["header"])
            ws.cell(row=row, column=4, value="; ".join(product["big_images"]))
        
        # Автоподбор ширины колонок
        for col in range(1, len(headers) + 1):
            max_length = 0
            column = get_column_letter(col)
            
            for row in range(1, len(products) + 2):
                try:
                    cell_value = str(ws[f'{column}{row}'].value)
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except:
                    pass
            
            adjusted_width = min((max_length + 2) * 1.2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Добавляем информацию о дате формирования отчета
        ws.cell(row=len(products) + 3, column=1, value=f"Отчет сформирован: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        ws.cell(row=len(products) + 4, column=1, value=f"Всего товаров с большими изображениями: {len(products)}")
        
        # Сохраняем файл
        filename = f"products_big_images_ASYNC_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        
        print(f"\n✅ Результат сохранен в Excel файл: {filename}")
        print(f"📊 Файл содержит {len(products)} товаров с большими изображениями")
        
        return len(products)

async def main_async():
    checker = AsyncBigImageChecker()
    
    try:
        print("🔐 Авторизация...")
        checker.authenticate()
        print("✅ Авторизация успешна")
        
        print("🚀 Асинхронная проверка размеров изображений...")
        products = await checker.get_products_with_big_images()
        
        print(f"\n📊 Итого найдено: {len(products)} товаров с изображениями > 750px")
        print(f"💾 Кэш изображений: {len(checker.image_cache)} уникальных файлов")
        
        if products:
            checker.save_to_excel(products)
        else:
            print("Товары с большими изображениями не найдены")
        
        return True
        
    except Exception as e:
        print(f"❌ Ошибка: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def main():
    return asyncio.run(main_async())

if __name__ == "__main__":
    main()
